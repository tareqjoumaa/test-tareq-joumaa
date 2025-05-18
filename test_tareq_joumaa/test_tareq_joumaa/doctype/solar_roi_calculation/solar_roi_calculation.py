import frappe
from frappe.model.document import Document
from frappe.utils.file_manager import get_file_path
from openpyxl import load_workbook
from datetime import datetime


class SolarROICalculation(Document):
    def before_save(self):
        if self.raw_data_file:
            data = self.extract_data_from_file(self.raw_data_file)
            if not data:
                frappe.throw("No valid data found in the uploaded file.")

            self.calculate_averages(data)
            self.calculate_monthly_tariffs(data)
    
        
    def extract_data_from_file(self, file_url):
        file_path = get_file_path(file_url)
        wb = load_workbook(filename=file_path, data_only=True)
        ws = wb.active

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            timestamp_idx = headers.index("Timestamp")
            kw_idx = headers.index("KW")
            kwh_idx = headers.index("KWH")
        except ValueError:
            frappe.throw("The Excel file must include 'Timestamp', 'KW', and 'KWH' columns.")

        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                timestamp = datetime.strptime(str(row[timestamp_idx]), "%Y-%m-%d %H:%M:%S")
            except Exception:
                frappe.throw(f"Invalid timestamp format in row: {row}")
            kw = row[kw_idx]
            kwh = row[kwh_idx]
            if kw is not None and kwh is not None:
                data.append({
                    "timestamp": timestamp,
                    "hour": timestamp.hour,
                    "month": timestamp.month,
                    "kw": kw,
                    "kwh": kwh,
                })
        return data
    

    def calculate_averages(self, data):
        self.average_kw = sum(d["kw"] for d in data) / len(data)
        self.average_kwh = sum(d["kwh"] for d in data) / len(data)


    def calculate_monthly_tariffs(self, data):
            self.set("monthly_usage", [])
            monthly_data = {}
            for d in data:
                monthly_data.setdefault(d["month"], []).append(d)

            for month in sorted(monthly_data.keys()):
                month_entries = monthly_data[month]
                low_kwh = [d["kwh"] for d in month_entries if d["hour"] >= 23 or d["hour"] < 6]
                high_kwh = [d["kwh"] for d in month_entries if 6 <= d["hour"] < 23]

                avg_low = 0.1 * (sum(low_kwh) / len(low_kwh)) if low_kwh else 0
                avg_high = 0.3 * (sum(high_kwh) / len(high_kwh)) if high_kwh else 0

                self.append("monthly_usage", {
                    "month": datetime(2023, month, 1).strftime('%B'),
                    "avg_low_tariff": avg_low,
                    "avg_high_tariff": avg_high
                })